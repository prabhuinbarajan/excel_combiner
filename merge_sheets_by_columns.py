import openpyxl
from merge_sheet_by_rows import *

from openpyxl.formula.translate import Translator






def get_idx(index, offset = 0, cl_type='row'):
    return index+offset if cl_type == 'row' else openpyxl.utils.cell.get_column_letter(index)
    pass


def get_group_ranges (rng, dimensions, offset=0, cl_type='row'):
    groupRange = None
    groupList = []
    if cl_type == 'row':
        for j in rng:
            if not dimensions[j].hidden:
                if groupRange is not None:
                    groupList.append(groupRange)

                groupRange = {'start': get_idx(j, offset, cl_type=cl_type), 'end': get_idx(j, offset,  cl_type=cl_type)}
            else:
                if groupRange is None:
                    groupRange = {'start': get_idx(j, offset, cl_type=cl_type)}
                groupRange['end'] = get_idx(j, offset, cl_type=cl_type)
        if groupRange is not None and groupRange['start'] != groupRange['end']:
            groupList.append(groupRange)
    else:
        for index in rng:
            dim  = dimensions.get(openpyxl.utils.cell.get_column_letter(index))
            if dim:
                groupList.append({'start': openpyxl.utils.cell.get_column_letter(dim.min+offset),
                                  'end': openpyxl.utils.cell.get_column_letter(dim.max+offset),
                                  'hidden' : dim.hidden} )
    return groupList


# Paste range
# Paste data from copyRange into template sheet
def apply_groups(dimensions, groupList, cl_type = 'row'):
    group_boundary = -1
    start = None
    end = None
    if cl_type == 'row':
        for gr in groupList:
            if gr['start'] != gr['end']:
                dimensions.group(start=gr['start'], end=gr['end']+group_boundary, hidden=True)
        #else:
        #    dimensions.group(start=gr['start'], end=gr['end'], hidden=gr['hidden'], outline_level=3)
    else:
        for i in range (0, len(groupList)):
            start = i
            while i< len(groupList) and groupList[i]['hidden']:
                i = i+1
            end = i-1
            if end > start:
                dimensions.group (start=groupList[start]['start'], end = groupList[end]['end'], hidden=True, outline_level=2)






def pasteRangeCols(startCol, startRow, endCol, endRow, sheetReceiving, sourceSheet,  sourceStartRowOffset=0, sourceStartColOffset=0, enableRowGrouping=False, enableColGrouping = False, extend_data=False, style_col = -1):
    countRow = 0+sourceStartRowOffset
    total_rows =  endRow-startRow
    total_cols = endCol-startCol
    groupRowList = get_group_ranges(range(sourceStartRowOffset, total_rows + sourceStartRowOffset), sourceSheet.row_dimensions, offset = startRow - sourceStartRowOffset, cl_type='row') \
        if enableRowGrouping else []
    groupColList = get_group_ranges(range(sourceStartColOffset+1, total_cols+ sourceStartColOffset+2), sourceSheet.column_dimensions, offset = startCol - sourceStartColOffset-1, cl_type='col') \
        if enableColGrouping else []

    for i in range(startRow, endRow + 1, 1):
        countCol = 1+ sourceStartColOffset
        for j in range(startCol, endCol + 1, 1):
            source = sourceSheet.cell(row=sourceStartRowOffset, column=j) \
                if extend_data else sourceSheet.cell(row=countRow+1, column=countCol)
            style_cell = sourceSheet.cell(row=countRow + 1, column=style_col) if style_col > 0 else source

            value = source.value
            target = sheetReceiving.cell(row=i, column=j)
            if source.data_type == 'f':
                target.value = Translator(value, source.coordinate).translate_formula(target.coordinate)
            else :
                target.value = value
            countCol += 1

            copy_style(style_cell, target)
        countRow += 1

    if enableRowGrouping:
        apply_groups ( sheetReceiving.row_dimensions, groupRowList, cl_type='row')
    if enableColGrouping:
        apply_groups(sheetReceiving.column_dimensions, groupColList, cl_type='col')



def mergeColumns(source_workbook, target_worksheet, regex=r".*", sourceStartRowOffset=0, sourceStartColOffset=0,
                 startRowOffset=0, startColOffset=0,  sourceEndColOffset=0, sourceEndRowOffset=0, columnGap = 0,
                 enableRowGrouping=False, enableColGrouping = False, applySourceOffsetFromFirst=False):
    print("Processing...")
    startCol = 1+startColOffset
    first = not applySourceOffsetFromFirst
    #enableRowGrouping = (not applySourceOffsetFromFirst) if enableRowGrouping else enableRowGrouping
    selector_regex = re.compile(regex)
    itemList = list(filter(lambda i: selector_regex.match(i), source_workbook.sheetnames))

    for sn in itemList:
        if sn == 'TOC':
            continue
        source_worksheet = source_workbook[sn]  # Add Sheet name
        startRow = 1+startRowOffset

        ssColOffset = 0 if first else sourceStartColOffset
        first = False

        endCol = startCol + (source_worksheet.max_column -1 - sourceEndColOffset-ssColOffset)
        endRow = source_worksheet.max_row - sourceEndRowOffset-sourceStartRowOffset+startRow
        print('sc:{} sr: {} ec: {} er: {} sn: {} '.format(startCol, startRow, endCol, endRow, sn))

        pasteRangeCols(startCol, startRow, endCol, endRow,
                                  target_worksheet, source_worksheet,
                       sourceStartRowOffset=sourceStartRowOffset, sourceStartColOffset=ssColOffset,
                       enableRowGrouping=enableRowGrouping, enableColGrouping=enableColGrouping)
        #enableRowGrouping = False
        startCol = endCol + 1 + columnGap
    # You can save the template as another file to create a new file here too.s
    print("Range transferred!")

def copy_data_in_range(worksheet=None, reference_row=None, col_rng = None, row_range=None, copy_format_column=-1):
    pasteRangeCols(col_rng.start, row_range.start, col_rng.stop, row_range.stop,
                   worksheet, worksheet,
                   sourceStartRowOffset=reference_row[0].row, extend_data = True,
                   style_col=copy_format_column)
