import openpyxl
from openpyxl.formula.translate import Translator
import re
from copy import copy


# Prepare the spreadsheets to copy from and paste too.

# File to be copied
workbook_base = 'Intl Vol Rate MTD QTD'

mtd_workbook = '{} P4 2020'.format(workbook_base)
ytd_workbook = 'Intl Vol Rate YTD P4 2020'
mtd_workbook_url = 'report_samples/{}.xlsx'.format(mtd_workbook)
ytd_workbook_url = 'report_samples/{}.xlsx'.format(ytd_workbook)

workbook_template = 'templates/{}.xltx' .format(workbook_base)
result_workbook = 'results/{}_combined.xlsx'.format(mtd_workbook)

mtd_qtd_wb = openpyxl.load_workbook(mtd_workbook_url)  # Add file name
ytd_wb = openpyxl.load_workbook(ytd_workbook_url)  # Add file name

mtd_regex = re.compile(r'^.*\-MTD')
qtd_regex = re.compile(r'^.*\-QTD')
ytd_regex = re.compile(r'^.*\-YTD')


# Paste range
# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData, rowOffset = 0 , colOffset = 0):
    countRow = 1
    for i in range(startRow, endRow , 1):
        countCol = 1
        for j in range(startCol, endCol , 1):
            source = copiedData.cell(row=countRow+ rowOffset, column=countCol)
            target = sheetReceiving.cell(row=i, column=j)
            #if type(cell).__name__ == 'MergedCell':

            if source.data_type == 'f':
                target.value = Translator(source.value, source.coordinate).translate_formula(target.coordinate)
            else :
                target.value = source.value
            #if source.has_style:
            #    target._style = copy(source._style)
            if source.has_style:
                target.font = copy(source.font)
                target.border = copy(source.border)
                target.fill = copy(source.fill)
                target.number_format = copy(source.number_format)
                target.protection = copy(source.protection)
                target.alignment = copy(source.alignment)

            countCol += 1
        countRow += 1


def createMergedSheet(worksheet, regex_filter, workbook):

    print("Processing...")
    startCol = 1
    startRow = 10
    initialRowOffset = 9

    itemList = list(filter(lambda i: regex_filter.match(i), workbook.sheetnames))
    for sn in itemList:
        sheet1 = workbook[sn]  # Add Sheet name
        endCol = sheet1.max_column
        endRow = startRow+ sheet1.max_row-17
        print('sc:{} sr: {} ec: {} er: {} sn: {}'.format(startCol, startRow, endCol, endRow,  sn))
        pasteRange(startCol, startRow, endCol, endRow,
                                  worksheet, sheet1, initialRowOffset)  # Change the 4 number values
        startRow=endRow

    # You can save the template as another file to create a new file here too.s
    print("Range copied and pasted!")


target = openpyxl.load_workbook(workbook_template)
target.template = False
#target = copy(wb_template)
createMergedSheet(target['YTD'], ytd_regex, ytd_wb)

createMergedSheet(target['MTD'], mtd_regex, mtd_qtd_wb)
createMergedSheet(target['QTD'], qtd_regex, mtd_qtd_wb)


target.save(result_workbook)
