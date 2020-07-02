import openpyxl
from openpyxl.formula.translate import Translator
import re
from copy import copy
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font


defaultFill = PatternFill("solid", fgColor="DDDDDD")

thick_top_border = Border(top=Side(style='thick'))
bold_font = Font(bold=True)

def apply_grand_total(grandtotalRows, worksheet, endRow , startCol , endCol, referenceRow,  grandTotalTitle = "Total"):
    row = worksheet[endRow]
    row[0].value = grandTotalTitle
    for j in range(startCol, endCol):
        target = row[j]
        source = referenceRow[j]
        if source.data_type == 'f':
            target.value = Translator(source.value, source.coordinate).translate_formula(target.coordinate)
        else:
            sumFormula = ''
            for sumRow in grandtotalRows:
                sumFormula = sumRow[j].coordinate if sumFormula == '' else sumFormula + '+' + sumRow[j].coordinate
            target.value = "=SUM(" + sumFormula + ")"
        copy_style(source, target)
    apply_style([row])


def copy_style (source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = copy(source.number_format)
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def add_separator(sheetReceiving, startCol, row, endCol, fill=defaultFill):
    for j in range(startCol, endCol, 1):
        target = sheetReceiving.cell(row=row, column=j)
        target.fill = copy(fill)

def apply_style(rows, border=thick_top_border, font=bold_font):
    for row in rows:
        for cell in row:
            if border:
                cell.border = border
            if font:
                cell.font = font

# Paste range
# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData, rowOffset = 0 , colOffset = 0):
    firstList = []
    lastList = []
    #typeList = []
    countRow = 1
    if startRow == endRow :
        endRow = startRow +1
    for i in list(range(startRow, endRow , 1)):
        countCol = 1
        for j in range(startCol, endCol , 1):

            source = copiedData.cell(row=countRow+ rowOffset, column=countCol)
            target = sheetReceiving.cell(row=i, column=j)
            #if type(target).__name__ == 'MergedCell':
            #    print('hello')
            if source.data_type == 'f':
                target.value = Translator(source.value, source.coordinate).translate_formula(target.coordinate)
            else :
                target.value = source.value
            #if source.has_style:
            #    target._style = copy(source._style)
            copy_style(source, target)

            if i == startRow:
                firstList.append(target.coordinate)
                #typeList.append(source.data_type)
            if i == endRow-1:
                lastList.append(target.coordinate)
            countCol += 1
        countRow += 1
    #return map(lambda x, y: str(x) + ":" + y, firstList, lastList)
    return [str(i) + ":" + str(j) for i, j in zip(firstList, lastList)] #, typeList


def createMergedSheet(worksheet, regex_filter, workbook, startCol, startRow, initialRowOffset, postRowShrinkage, groupRows=False, subtotalRows=False, totalColOffset = 0, totalColOffsetUpperBound = -1, subtotalFunctionNum = 9,   grandTotal = False, grandTotalTitle=None ):

    print("Processing...")
    itemList = list(filter(lambda i: regex_filter.match(i), workbook.sheetnames))
    firstSheet = None
    subtotalOffset = 1 if subtotalRows else 0
    listOfSubTotals = []

    for sn in itemList:
        startRow += subtotalOffset
        sheet1 = workbook[sn]  # Add Sheet name
        firstSheet = sheet1 if sheet1 == None else sheet1
        endCol = sheet1.max_column
        endRow = startRow+ sheet1.max_row-initialRowOffset-postRowShrinkage
        print('sc:{} sr: {} ec: {} er: {} sn: {} subtotal {} grandTotal {}'.format(startCol, startRow, endCol, endRow,  sn, subtotalRows, grandTotal))
        subtotalCoordinates  = pasteRange(startCol, startRow, endCol, endRow,
                                  worksheet, sheet1, initialRowOffset)  # Change the 4 number values

        if subtotalRows:
            subTotalTitle = sn.replace("-MTD","").replace("-YTD","").replace("-QTD","")
            worksheet.cell(row=startRow - 1, column=2).value = subTotalTitle
            listRowSubTotal = []
            for j in range(totalColOffset, endCol, 1):
                target = worksheet.cell(row=startRow-1, column=j)
                source = worksheet.cell(row=startRow, column=j)
                if source.data_type == 'f':
                    target.value = Translator(source.value, source.coordinate).translate_formula(target.coordinate)
                else:
                    target.value = "=SUBTOTAL(" + str(subtotalFunctionNum) + "," + subtotalCoordinates[j-1] + ")"
                copy_style(source, target)
                listRowSubTotal.append(target.coordinate)
            if grandTotal:
                listOfSubTotals.append(listRowSubTotal)
            rows = [worksheet[startRow - 1]]
            apply_style(rows, border=None)

            if len(itemList) == 1:
                if grandTotalTitle:
                    worksheet.cell(row=startRow - 1, column=1).value = grandTotalTitle

        if groupRows:
            #for idx in range(startRow, endRow):
            worksheet.row_dimensions.group(start=startRow, end=endRow-1, hidden= True)

        startRow=endRow

        #break
    if grandTotal:
        startCell = worksheet.cell(row=endRow, column=1)
        startCell.value = "Total" if grandTotalTitle is None else grandTotalTitle
        grandTotalList = []
        if subtotalRows:
            temp = []
            grandTotalList = ['' for i in listOfSubTotals[0]]
            for item in listOfSubTotals:
                temp = [str(i) + "+" + str(j) if len(i) > 0 else j for i, j in zip(grandTotalList, item)]
                grandTotalList = temp
        else:
            for j in range (startCol , endCol):
                grandTotalList.append(worksheet.cell(startRow, j).coordinate + ":" + worksheet.cell(endRow , j).coordinate)
        if totalColOffsetUpperBound > 0:
            endCol = totalColOffsetUpperBound
        for j in range(totalColOffset, endCol, 1):
            target = worksheet.cell(row=endRow, column=j)
            source = worksheet.cell(row=endRow-2, column=j)
            if source.data_type == 'f':
                target.value = Translator(source.value, source.coordinate).translate_formula(target.coordinate)
            else:
                target.value = "=SUM(" + grandTotalList[j-totalColOffset] + ")"
            copy_style(source, target)

        rows = [worksheet[endRow]]
        apply_style(rows)



    # You can save the template as another file to create a new file here too.s
    print("Range copied and pasted!")
    return firstSheet, endRow
