import openpyxl
from config_reader import *
import pandas as pd
import re
from openpyxl.utils.dataframe import dataframe_to_rows


# Prepare the spreadsheets to copy from and paste too.
(TB_input_path,PL_input_path,template_path,TB_output_path,PL_output_path,myyear,myper) = get_config(env=sys.argv[1] if len(sys.argv) > 1 else None)

# File to be copied



def pasteRange(ws, rows, startRow,totalsFrom, title=''):
    range_text = ''
    lastRow = 1
    lastCol = 1
    first=True
    lr = None
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            lr = row
            ws.cell(row=r_idx+startRow, column=c_idx, value=value)
            lastRow = r_idx+startRow
            lastCol = c_idx
            if first:
                range_text=ws.cell(row=lastRow, column=lastCol).coordinate
                first=False

    if (totalsFrom >= 0):

        for c_idx, value in enumerate(lr, 1):
            if c_idx == 1 :
                ws.cell(row=lastRow + 1, column=c_idx, value=title)
            if c_idx > totalsFrom:
                column_letter =  ws.cell(row=startRow, column=c_idx).column_letter
                range_text = column_letter +str(startRow+1) + ":" + column_letter+str(lastRow)
                ws.cell(row=lastRow+1, column=c_idx,  value='=SUM('+range_text + ')')
    return lastRow+1, lr

pCY_workbook = fnmatch.filter(os.listdir(TB_input_path), '*A510G_2020*')
pPY_workbook = fnmatch.filter(os.listdir(TB_input_path), '*A510G_2019*')
workbook = fnmatch.filter(os.listdir(TB_input_path), '*A510-BSTNDLC Detail - by Account Category*')
workbook1 = pCY_workbook[0]
workbook2 = pPY_workbook[0]
workbook3 = workbook[0]
print("File Names are " + pCY_workbook[0] + ", " + pPY_workbook[0] + ", " + workbook[0])

pCY_data_file = r'{}{}'.format(TB_input_path,pCY_workbook[0])
pPY_data_file = r'{}{}'.format(TB_input_path,pPY_workbook[0])
workbook_url = r'{}{}'.format(TB_input_path,workbook[0])
account_meta = r'{}account_meta.csv'.format(TB_input_path)
print("Workbook URLs are " + pCY_data_file + ", " + pPY_data_file + ", " + workbook_url + ", " + account_meta)

interested_period = 'P4 2020'
period_end_date = '4/18/2020'
num_periods = 13

workbook_template = r'{}{}.xltx' .format(template_path,workbook)
print("Template URL is " + workbook_template)

result_workbook = r'{}{}_combined.xlsx'.format(TB_output_path,workbook)
print("Result workbook URLs are " + result_workbook)

period_columns_regex = re.compile(r'^P[0-9]*')

account_meta = pd.read_csv(account_meta, dtype=str)

df_PY = pd.read_csv(pPY_data_file)
df_PY= df_PY.loc[:, ~df_PY.columns.str.contains('^Unnamed')]
df_PY.drop(['Department', 'Company', 'Budget Entity'], axis=1, inplace=True)
df_PY['ac1'] = pd.to_numeric(df_PY['Account'], errors='coerce')
df_CY = pd.read_csv(pCY_data_file)
df_CY.drop(['Department', 'Company', 'Budget Entity'], axis=1, inplace=True)
df_CY= df_CY.loc[:, ~df_CY.columns.str.contains('^Unnamed')]
df_CY['ac1'] = pd.to_numeric(df_CY['Account'], errors='coerce')


df_PY['AffiliateRollup'] = df_PY['Affiliate'].str.slice(0, 4, 1)
df_CY['AffiliateRollup'] = df_CY['Affiliate'].str.slice(0, 4, 1)


def getCumulativePLYTD(df):

    df_col_l = list(filter(lambda i: period_columns_regex.match(i), df.columns))
    df_period_columns = dict(zip(df_col_l, map(lambda x: 'sum', df_col_l)))
    df_agg_sum = df[(df['ac1'] >= 400000) & (df['ac1'] <= 999999)].agg(df_period_columns).cumsum()
    return df_agg_sum


df_PY_agg_sum = getCumulativePLYTD(df_PY)
df_CY_agg_sum = getCumulativePLYTD(df_CY)


df_net_rev_agg = pd.concat([df_PY_agg_sum , df_CY_agg_sum])
df_net_rev_agg ['Account'] = '9999999'
df_net_rev_agg ['AffiliateRollup'] = ''
df_net_rev_agg ['Affiliate'] = ''
df_net_rev_agg = df_net_rev_agg.to_frame().T

#result = pd.merge(df_PY, df_CY, how='outer', on=['Account', 'AffiliateRollup', 'Affiliate'])
result = pd.concat([df_PY, df_CY,df_net_rev_agg])
result.drop(['Affiliate'], axis=1, inplace=True)
result.rename(columns = {'AffiliateRollup':'Affiliate'}, inplace = True)

columns = list(filter(lambda i: period_columns_regex.match(i), result.columns))

index_of_period= columns.index(interested_period)
columns = columns[index_of_period-num_periods:index_of_period+1]
aggregate_columns = {columns[i]: 'sum' for i in range(0, len(columns))}


result = pd.merge(result, account_meta, how='left', on=['Account'])

agg_result = result.groupby(['AC', 'Account', 'Affiliate']).agg(aggregate_columns).reset_index()
sub_totals = result.groupby(['AC']).agg(aggregate_columns).reset_index()



check_figure = sub_totals.agg(['sum'])

combined_result = pd.merge(agg_result, account_meta, how='outer', on=['AC', 'Account'])
enriched_trend_cols = ['AC', 'Account Category', 'Account', 'Acct Cat', 'Affiliate', 'Account Description']
enriched_trend_cols.extend(columns)

for col in ['Account Category', 'Account', 'Acct Cat', 'Affiliate', 'Account Description']:
    check_figure[col] = ''
    sub_totals[col] = ''

final_result = combined_result.reindex(enriched_trend_cols,axis=1)
sub_totals = sub_totals.reindex(enriched_trend_cols,axis=1)
check_figure = check_figure.reindex(enriched_trend_cols,axis=1)
check_figure = check_figure.drop(['AC'], axis=1)


target = openpyxl.load_workbook(workbook_template)
target.template = False
sheet = target['Sheet']

column_header = pd.DataFrame(columns=final_result.columns)
column_header.drop(['AC'], axis = 1, inplace=True)
header_rows = dataframe_to_rows(column_header, index=False, header=True)
pasteRange(sheet, header_rows, 8,-1)

startRow = 9
rows = None
totalRows = {}
lastRowRecord = None
for asset_category in ['Assets', 'Liabilities','Equities']:
    ac_results = final_result[final_result['AC'] == asset_category]
    ac_results = ac_results.drop(['AC'], axis=1)
    rows = dataframe_to_rows(ac_results, index=False, header=False)
    lr , lastRowRecord = pasteRange(sheet, rows, startRow, 5, 'Total '+ asset_category)
    totalRows[asset_category] = lr
    startRow += ac_results.shape[0]+1


for c_idx, value in enumerate(lastRowRecord,1):
    if c_idx == 1 :
        sheet.cell(row=startRow + 2, column=c_idx, value='check figure')

    if c_idx > 5:
        column_letter = sheet.cell(row=startRow+2, column=c_idx).column_letter
        range_text = column_letter + str(totalRows['Assets']) + \
                     '+' + column_letter + str(totalRows['Liabilities']) + \
                     '+' + column_letter + str(totalRows['Equities'])
        sheet.cell(row=startRow+2, column=c_idx,  value='=' + range_text )
sheet['A7'] = period_end_date
target.save(result_workbook)
print(agg_result)


